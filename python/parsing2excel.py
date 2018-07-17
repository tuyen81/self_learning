#!/usr/bin/env python
# -*- coding: utf-8 -*-

from openpyxl import Workbook
import os
import os.path
import re
import sys
import shutil

# PATH_TO_PROJECT = os.getcwd()
input_data_file = "compare_cip_bxt_with_cip_diff.txt"

def split_output_per_commit (input_data_file):
    """ We want to saparate information of each commit to a file. Parse line to line
    and when a line start with commit, we will open a new temple file for storing
    the information of the commit with the name of file is $commit_id.log
    """
    try:
        commit_all = open(input_data_file)
    except IOError:
        print 'Cannot open the "%s" file' % input_data_file
        sys.exit(1)

    if os.path.exists('tmp'):
        shutil.rmtree('tmp')
    try:
        os.makedirs('tmp')
    except OSError:
        print 'Error: Cannot create the tmp directory'

    lines = commit_all.readlines()
    commit_all.close()
    # The begin of the input file without start with commit string often is comments
    # for how to create the input file and they are not date about commits which
    # we want. So, store them to the comment.txt file.
    commit_each = open("comment.txt", "w")
    commits = []
    for line in lines:
        m = re.compile("^commit .*").match(line)
        if m is not None:
            loop_end = 0
            in_loop = 1
            commit = m.group(0).split()[1]
            commits.append(commit)
            try:
                commit_each = open("./tmp/%s.log" % commit, "w")
            except IOError:
                print '"./tmp/%s.log" cannot be opened' % commit
                sys.exit(1)

        # m = re.compile("^Author: .*").match(line)
        # if m is not None:
        #    author = m.group(0).split(' ', 1)[1]
        commit_each.write("%s" % line)

    return commits

def read_author (commit):
    try:
        commit_each = open("./tmp/%s.log" % commit)
    except IOError:
        print '"./tmp/%s.log" cannot be opened' % commit
        sys.exit(1)

    data = commit_each.read()
    commit_each.close()

    m = re.compile("^Author: (.*)", re.M).search(data)
    if m is not None:
        author = m.group(1)
    else:
        author = ""

    return author

def read_date (commit):
    try:
        commit_each = open("./tmp/%s.log" % commit)
    except IOError:
        print '"./tmp/%s.log" cannot be opened' % commit
        sys.exit(1)

    data = commit_each.read()
    commit_each.close

    m = re.compile("^Date: (.*)", re.M).search(data)
    if m is not None:
        date = m.group(1)
    else:
        date = ""

    return date

def read_message (commit):
    try:
        commit_each = open("./tmp/%s.log" % commit)
    except IOError:
        print '"./tmp/%s.log" cannot be opened'
        sys.exit(1)

    data = commit_each.read()
    commit_each.close

    m = re.compile("^ (.*)", re.M | re.S).search(data)
    if m is not None:
        message = m.group(0)
    else:
        message = ""

    return message

# This section is for formatting the excel file
# Open a workbook
wb = Workbook()
sheet = wb.create_sheet(title="commit_info")

commits = split_output_per_commit (input_data_file)

headers = ["Commit ID", "Author", "Date", "Commit message"]

i = 1
for j in range (1, 5):
    sheet.cell(row = i, column = j).value = headers[j - 1]

for commit in commits:
    author = read_author(commit)
    date = read_date(commit)
    message = read_message(commit)

    infor = [commit, author, date, message]
    i = i + 1
    for j in range (1, 5):
        sheet.cell(row = i, column = j).value = infor[j - 1]

# if we have added sheets, remove the default one ("Sheet")
for i in wb.worksheets:
    if i.title == 'Sheet':
        wb.remove(i)

wb.save("commit_info.xlsx")
